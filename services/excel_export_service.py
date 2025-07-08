import logging
from datetime import datetime
from io import BytesIO
from typing import Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services import RegionalDataService

logger = logging.getLogger(__name__)


class ExcelExportService:
    """Service class for Excel export functionality."""

    def __init__(self):
        self.regional_service = RegionalDataService()

    def create_excel_export(self, parsed_date: str) -> Tuple[BytesIO, str]:
        """
        Create Excel export for regional COVID-19 data.

        Args:
            parsed_date: Date to export data for ('latest' or date string)

        Returns:
            Tuple of (excel_buffer, filename)

        Raises:
            ValueError: If no data available for the specified date
        """

        regional_summaries = self.regional_service.get_regional_summary_for_date(
            target_date=parsed_date, sort_by="cases_desc"
        )

        if not regional_summaries:
            raise ValueError("No data available for the specified date")

        wb = Workbook()
        ws = wb.active
        ws.title = "COVID-19 Regional Data"

        self._populate_excel_worksheet(ws, regional_summaries, parsed_date)

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        filename = self._generate_filename(parsed_date)

        logger.info(f"Excel export completed: {filename}")
        return excel_buffer, filename

    def _populate_excel_worksheet(self, ws, regional_summaries, parsed_date):
        """Populate the Excel worksheet with data and styling."""

        styles = self._get_excel_styles()

        date_str = str(parsed_date) if parsed_date != "latest" else "Latest"
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value = f"COVID-19 Regional Data - {date_str}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws["A3"] = "Region"
        ws["B3"] = "Total Cases"

        for col in ["A3", "B3"]:
            cell = ws[col]
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["header_alignment"]
            cell.border = styles["header_border"]

        row = 4
        for summary in regional_summaries:
            ws[f"A{row}"] = summary.region_name
            ws[f"B{row}"] = summary.total_cases

            ws[f"A{row}"].alignment = styles["cell_alignment"]
            ws[f"A{row}"].border = styles["cell_border"]
            ws[f"B{row}"].alignment = styles["number_alignment"]
            ws[f"B{row}"].border = styles["cell_border"]

            row += 1

        total_cases = sum(summary.total_cases for summary in regional_summaries)
        summary_row = row + 1

        ws[f"A{summary_row}"] = "TOTAL"
        ws[f"B{summary_row}"] = total_cases

        summary_font = Font(bold=True)
        ws[f"A{summary_row}"].font = summary_font
        ws[f"B{summary_row}"].font = summary_font
        ws[f"A{summary_row}"].alignment = styles["cell_alignment"]
        ws[f"B{summary_row}"].alignment = styles["number_alignment"]
        ws[f"A{summary_row}"].border = styles["header_border"]
        ws[f"B{summary_row}"].border = styles["header_border"]

        self._add_metadata(ws, summary_row + 3, date_str, len(regional_summaries))

        self._adjust_column_widths(ws)

    def _get_excel_styles(self) -> dict:
        """Get Excel styling definitions."""
        return {
            "header_font": Font(bold=True, color="FFFFFF"),
            "header_fill": PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            ),
            "header_alignment": Alignment(horizontal="center", vertical="center"),
            "header_border": Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            ),
            "cell_alignment": Alignment(horizontal="left", vertical="center"),
            "number_alignment": Alignment(horizontal="right", vertical="center"),
            "cell_border": Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            ),
        }

    def _add_metadata(self, ws, metadata_row: int, date_str: str, total_regions: int):
        """Add metadata information to the worksheet."""
        ws[f"A{metadata_row}"] = "Export Date:"
        ws[f"B{metadata_row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws[f"A{metadata_row + 1}"] = "Data Date:"
        ws[f"B{metadata_row + 1}"] = date_str
        ws[f"A{metadata_row + 2}"] = "Total Regions:"
        ws[f"B{metadata_row + 2}"] = total_regions

    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except BaseException:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _generate_filename(self, parsed_date: str) -> str:
        """Generate appropriate filename for the export."""
        if parsed_date == "latest":
            return "covid19_italy_regional_data_latest.xlsx"
        else:
            return f"covid19_italy_regional_data_{parsed_date}.xlsx"
